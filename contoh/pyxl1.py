# tambahkan catatan pada bagian yang ditandai #

from openpyxl import Workbook, load_workbook

# membaca file excel
wb = load_workbook("Book1.xlsx")

# baca worksheet
ws = wb["Sheet1"]

# 
print(ws["A2"].value)

# 
print(ws.cell(row=1, column=2).value)