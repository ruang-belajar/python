# tambahkan catatan pada bagian yang ditandai #


from random import randint
from openpyxl import Workbook, load_workbook

#
wb = Workbook()

#
ws = wb.active

#
ws.cell(row=1, column=1).value = "100 data acak"

#
n = 1
while n<=100:
    ws.cell(row=2+n, column=1).value = randint(1,100)
    n += 1

#
wb.save("contoh1.xlsx")