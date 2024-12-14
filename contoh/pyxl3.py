# tambahkan catatan pada bagian yang ditandai #

from random import randint
from openpyxl import Workbook, load_workbook

# 
wb = Workbook()

# 
ws = wb.active

# 
ws.cell(row=1, column=1).value = randint(10,100)

#
n = 1
while n<=5:
    wb.save("file"+str(n)+".xlsx")
    n += 1